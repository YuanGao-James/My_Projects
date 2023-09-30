import os
import openpyxl
import numpy as np
from PIL import Image
import random
import cv2
import matplotlib.pyplot as plt
import math
from scipy import ndimage

# names: ['mcp2','pip2','dip2','mcp3','pip3','dip3','mcp4','pip4','dip4','mcp5','pip5','dip5']

# pixel distance
'''
ground_truth_dir = 'C:/代码/BU/hand_seg/YuanGao/yolov5-master/My data/train2_redivide_and_cleaning/labels/test'
prediction_dir = 'C:/代码/BU/hand_seg/YuanGao/yolov5-master/runs/detect/test3/labels'
img_dir = 'C:/代码/BU/hand_seg/YuanGao/yolov5-master/My data/train2_redivide_and_cleaning/images/test'

# excel load
wb = openpyxl.load_workbook('D:/高源/BU/AICV Lab/Hand_OA/Voxel_size.xlsx')
ws = wb.active
title = ['mcp2','pip2','dip2','mcp3','pip3','dip3','mcp4','pip4','dip4','mcp5','pip5','dip5']
for i in range(len(title)):
    ws.cell(row=1, column=i + 13).value = title[i]
pointer = 2
for filename in os.listdir(prediction_dir):
    filename = filename[:-4]
    print(filename)
    while 1:
        if int(ws.cell(row=pointer, column=1).value) != int(filename[:7]):
            pointer += 1
        else:
            break

    image = Image.open(img_dir+'/'+filename+'.png')
    (x, y) = image.size

    with open(ground_truth_dir + '/' + filename + '.txt', "r") as f1:
        ground_truth = f1.readlines()
        with open(prediction_dir + '/' + filename + '.txt', "r") as f2:
            prediction = f2.readlines()
            pre_box = np.zeros((12, 3))
            for p in prediction:
                p = p.split()
                if float(p[5]) > pre_box[int(p[0])][2]:
                    pre_box[int(p[0])][0] = float(p[1])
                    pre_box[int(p[0])][1] = float(p[2])
                    pre_box[int(p[0])][2] = float(p[5])
            for i in range(12):
                g = ground_truth[i].split()
                if pre_box[i][2] == 0:
                    continue
                else:
                    dx = (float(g[1]) - pre_box[i][0]) * x
                    dy = (float(g[2]) - pre_box[i][1]) * y
                    distance = np.sqrt(dx**2 + dy**2)
                    ws.cell(row=pointer, column=i + 13).value = distance

wb.save('D:/高源/BU/AICV Lab/Hand_OA/Voxel_size3.xlsx')
'''


# Visualize the distance by points instead of boxes
'''
ground_truth_dir = 'C:/代码/BU/hand_seg/YuanGao/yolov5-master/My data/train2_redivide_and_cleaning/labels/train'
prediction_dir = 'C:/代码/BU/hand_seg/YuanGao/yolov5-master/runs/detect/train3/labels'
img_dir = 'C:/代码/BU/hand_seg/YuanGao/yolov5-master/My data/train2_redivide_and_cleaning/images/train'

filename = '9001695'
filename = filename + '_v06'
img = cv2.imdecode(np.fromfile(img_dir + '/' + filename + '.png', dtype=np.uint8), 1)
(x, y, _) = img.shape

with open(ground_truth_dir + '/' + filename + '.txt', "r") as f1:
    ground_truth = f1.readlines()
    truth_box = np.zeros((12, 2))
    for g in ground_truth:
        g = g.split()
        truth_box[int(g[0])][0] = int(float(g[1]) * y)
        truth_box[int(g[0])][1] = int(float(g[2]) * x)
    with open(prediction_dir + '/' + filename + '.txt', "r") as f2:
        prediction = f2.readlines()
        pre_box = np.zeros((12, 3))
        for p in prediction:
            p = p.split()
            if float(p[5]) > pre_box[int(p[0])][2]:
                pre_box[int(p[0])][0] = int(float(p[1]) * y)
                pre_box[int(p[0])][1] = int(float(p[2]) * x)
                pre_box[int(p[0])][2] = float(p[5])

img[int(pre_box[10][1])-3:int(pre_box[10][1])+3, int(pre_box[10][0])-3:int(pre_box[10][0])+3, 1] = 0
img[int(truth_box[10][1])-3:int(truth_box[10][1])+3, int(truth_box[10][0])-3:int(truth_box[10][0])+3, 0] = 0

plt.imshow(img)
plt.show()
'''


# data_cut
'''
img_dir = 'D:/BU_CV/Hand_OA/3557 image+label'
label_dir = 'D:/BU_CV/Hand_OA/new_label'
goal_img_dir = 'C:/代码/BU/hand_seg/YuanGao/yolov5-master/My data/train2_redivide_training_set'
test = [0,0,0]
for filename in os.listdir(label_dir):
    filename = filename[:-4]
    print(filename)
    image = Image.open(img_dir + '/' + filename + '.png')
    r = random.random()
    if r < 0.7:
        image.save(goal_img_dir + '/images/train/' + filename + '.png')
        fw = open(goal_img_dir + '/labels/train/' + filename + '.txt', 'w')  # 将要输出保存的文件地址
        with open(label_dir + '/' + filename + '.txt', "r") as f:
            data = f.readlines()
            for d in data:
                fw.write(d)

    elif r < 0.85:
        image.save(goal_img_dir + '/images/val/' + filename + '.png')
        fw = open(goal_img_dir + '/labels/val/' + filename + '.txt', 'w')  # 将要输出保存的文件地址
        with open(label_dir + '/' + filename + '.txt', "r") as f:
            data = f.readlines()
            for d in data:
                fw.write(d)

    else:
        image.save(goal_img_dir + '/images/test/' + filename + '.png')
        fw = open(goal_img_dir + '/labels/test/' + filename + '.txt', 'w')  # 将要输出保存的文件地址
        with open(label_dir + '/' + filename + '.txt', "r") as f:
            data = f.readlines()
            for d in data:
                fw.write(d)
'''


# txt_refine
'''
dir_input = 'D:/BU_CV/Hand_OA/3557 image+label'
dir_output = 'D:/BU_CV/Hand_OA/new_label'

for filename in os.listdir(dir_input):
    if filename[-3:] != 'txt':
        continue

    image = Image.open(dir_input + '/' + filename[:-3] + 'png')
    (x, y) = image.size

    fw = open(dir_output + '/' + filename, 'w')  # 将要输出保存的文件地址
    with open(dir_input + '/' + filename, "r") as f:
        data = f.readlines()
        del(data[0])
        for i in range(12):
            line = data[i]
            line_split = line.split()

            num1 = int(line_split[1])/x
            num2 = int(line_split[2])/y
            num3 = 180/x
            num4 = 180/y
            line_input = str(i) + ' ' + str(num1) + ' ' + str(num2) + ' ' + str(num3) + ' ' + str(num4)
            print(line_input)
            fw.write(line_input+'\n')
'''


# box with angle
'''
prediction_dir = 'C:/代码/BU/hand_seg/YuanGao/yolov5-master/runs/detect/all3/labels'
img_dir = 'D:/高源\BU/AICV Lab/Hand_OA/3557 image+label'
save_dir = 'C:/代码/BU/hand_seg/YuanGao/yolov5-master/My data/result_with_angle'
box_size = (150, 200)

# for filename in os.listdir(img_dir):
#     filename = filename[:-4]
filename = '9022408_v06'
img = cv2.imdecode(np.fromfile(img_dir + '/' + filename + '.png', dtype=np.uint8), -1)
(x, y) = img.shape

with open(prediction_dir + '/' + filename + '.txt', "r") as f2:
    prediction = f2.readlines()
    pre_box = np.zeros((12, 3))
    angle_box = np.zeros((4, 2))

    # extract txt to pre_box
    for p in prediction:
        p = p.split()
        if float(p[5]) > pre_box[int(p[0])][2]:
            pre_box[int(p[0])][0] = int(float(p[1]) * y)
            pre_box[int(p[0])][1] = int(float(p[2]) * x)
            pre_box[int(p[0])][2] = float(p[5])
    # calculate angle
    for i in range(4):
        angle1 = np.arctan((pre_box[3*i+1][0]-pre_box[3*i+2][0])/(pre_box[3*i+2][1]-pre_box[3*i+1][1]))/np.pi*180
        angle2 = np.arctan((pre_box[3*i+0][0]-pre_box[3*i+1][0])/(pre_box[3*i+1][1]-pre_box[3*i+0][1]))/np.pi*180
        angle_box[i][0] = angle1
        angle_box[i][1] = angle2
    # draw box
    for i in range(4):
        rect = ((pre_box[3 * i + 0][0], pre_box[3 * i + 0][1]), box_size, angle_box[i][1])
        box = cv2.boxPoints(rect)
        box = np.int0(box)
        cv2.drawContours(img, [box], 0, (255, 255, 255), 3)

        rect = ((pre_box[3 * i + 1][0], pre_box[3 * i + 1][1]), box_size, (angle_box[i][0]+angle_box[i][1])/2)
        box = cv2.boxPoints(rect)
        box = np.int0(box)
        cv2.drawContours(img, [box], 0, (255, 255, 255), 5)

        rect = ((pre_box[3 * i + 2][0], pre_box[3 * i + 2][1]), box_size, angle_box[i][0])
        box = cv2.boxPoints(rect)
        box = np.int0(box)
        cv2.drawContours(img, [box], 0, (255, 255, 255), 8)
    # mark center
    for i in range(12):
        x = int(pre_box[i][0])
        y = int(pre_box[i][1])
        img[y-5:y+5, x-5:x+5] = 0

# cv2.imencode('.png', img)[1].tofile(save_dir+'/'+filename+'.png')
plt.imshow(img, 'gray')
plt.show()
'''


# crop with angle
'''
def crop(image, x, y, width, height):
    return image[y:y + height, x:x + width]

def center_crop(image, x, y, width, height):
    return crop(image, int(x - (width / 2)), int(y - (height / 2)), width, height)

def angled_center_crop(image, x, y, width, height, angle):
    # print(angle)
    # return center_crop(image, x, y, width, height)

    start_point = (int(x - (width * 0.5)), int(y - (height * 0.5)))
    end_point = (start_point[0] + width, start_point[1] + height)
    points = points4(start_point, end_point)


    rotated = rotate_points(points, (x, y), angle)

    min_x = min(rotated[0][0], rotated[1][0], rotated[2][0], rotated[3][0])
    max_x = max(rotated[0][0], rotated[1][0], rotated[2][0], rotated[3][0])

    min_y = min(rotated[0][1], rotated[1][1], rotated[2][1], rotated[3][1])
    max_y = max(rotated[0][1], rotated[1][1], rotated[2][1], rotated[3][1])

    # crop the image around the super bounds
    temp = crop(image, min_x, min_y, max_x - min_x, max_y - min_y)

    # cv2_imshow(temp)

    degrees = math.degrees(angle)
    temp = ndimage.rotate(temp, degrees)

    rotated_height = temp.shape[0]
    rotated_width = temp.shape[1]

    center_x = int(rotated_width * 0.5)
    center_y = int(rotated_height * 0.5)

    return center_crop(temp, center_x, center_y, width, height)

def points4(start_point, end_point):
    points = []
    points.append(start_point)
    points.append((end_point[0], start_point[1]))
    points.append(end_point)
    points.append((start_point[0], end_point[1]))
    return points

def z_rotate(origin, point, angle):
    ox, oy = origin
    px, py = point

    qx = ox + math.cos(angle) * (px - ox) - math.sin(angle) * (py - oy)
    qy = oy + math.sin(angle) * (px - ox) + math.cos(angle) * (py - oy)
    return int(qx), int(qy)

def rotate_points(points, center, angle):
    rotated = []
    n = len(points)
    for i in range(n):
        rotated.append(z_rotate(center, points[i], angle))
    return rotated

prediction_dir = 'C:/代码/BU/hand_seg/YuanGao/yolov5-master/runs/detect/all3/labels'
img_dir = 'D:/高源/BU/AICV Lab/Hand_OA/3557 image+label'
save_dir = 'C:/代码/BU/hand_seg/YuanGao/yolov5-master/My data/crops_with_angle_all_2'
box_size = (150, 150)

for filename in os.listdir(prediction_dir):
    filename = filename[:-4]
    # filename = '9001695_v06'
    print(filename)
    img = cv2.imdecode(np.fromfile(img_dir + '/' + filename + '.png', dtype=np.uint8), -1)
    (x, y) = img.shape

    with open(prediction_dir + '/' + filename + '.txt', "r") as f2:
        prediction = f2.readlines()
        pre_box = np.zeros((12, 3))
        angle_box = np.zeros((4, 2))

        # extract txt to pre_box
        for p in prediction:
            p = p.split()
            if float(p[5]) > pre_box[int(p[0])][2]:
                pre_box[int(p[0])][0] = int(float(p[1]) * y)
                pre_box[int(p[0])][1] = int(float(p[2]) * x)
                pre_box[int(p[0])][2] = float(p[5])

        # calculate angle
        for i in range(4):
            angle1 = np.arctan((pre_box[3*i+1][0]-pre_box[3*i+2][0])/(pre_box[3*i+2][1]-pre_box[3*i+1][1]))
            angle2 = np.arctan((pre_box[3*i+0][0]-pre_box[3*i+1][0])/(pre_box[3*i+1][1]-pre_box[3*i+0][1]))
            angle_box[i][0] = angle1
            angle_box[i][1] = angle2

    crops = []
    for i in range(4):
        crops.append(angled_center_crop(img, pre_box[i*3][0], pre_box[i*3][1], box_size[0], box_size[1], angle_box[i][0]))
        crops.append(angled_center_crop(img, pre_box[i*3+1][0], pre_box[i*3+1][1], box_size[0], box_size[1], (angle_box[i][0]+angle_box[i][1])/2))
        crops.append(angled_center_crop(img, pre_box[i*3+2][0], pre_box[i*3+2][1], box_size[0], box_size[1], angle_box[i][1]))

    names = ['mcp2','pip2','dip2','mcp3','pip3','dip3','mcp4','pip4','dip4','mcp5','pip5','dip5']
    for i in range(12):
        if crops[i] != []:
            cv2.imencode('.png', crops[i])[1].tofile(save_dir + '/' + filename + '_' + names[i] + '.png')

    # plt.subplot(1, 2, 1)
    # plt.imshow(img, 'gray')
    # plt.subplot(1, 2, 2)
    # plt.imshow(crops[10], 'gray')
    # plt.show()
'''


# calculate error of angle
'''
prediction_dir = 'C:/代码/BU/hand_seg/YuanGao/yolov5-master/runs/detect/all3/labels'
ground_truth_dir = 'D:/高源/BU/AICV Lab/Hand_OA/3557 image+label'
img_dir = 'D:/高源/BU/AICV Lab/Hand_OA/3557 image+label'
box_size = (150, 150)

# excel load
wb = openpyxl.load_workbook('D:/高源/BU/AICV Lab/Hand_OA/Voxel_size.xlsx')
ws = wb.active
title = ['mcp2','pip2','dip2','mcp3','pip3','dip3','mcp4','pip4','dip4','mcp5','pip5','dip5']
for i in range(len(title)):
    ws.cell(row=1, column=i + 13).value = title[i]
pointer = 2

for filename in os.listdir(prediction_dir):
    filename = filename[:-4]
    print(filename)
    while 1:
        if int(ws.cell(row=pointer, column=1).value) != int(filename[:7]):
            pointer += 1
        else:
            break

    image = Image.open(img_dir+'/'+filename+'.png')
    (x, y) = image.size

    with open(ground_truth_dir + '/' + filename + '.txt', "r") as f1:
        ground_truth = f1.readlines()
        truth_box = np.zeros(12)
        for i in range(1, 13):
            g = ground_truth[i]
            g = g.split()
            truth_box[i-1] = float(g[3])
        with open(prediction_dir + '/' + filename + '.txt', "r") as f2:
            prediction = f2.readlines()
            pre_box = np.zeros((12, 3))

            # extract txt to pre_box
            for p in prediction:
                p = p.split()
                if float(p[5]) > pre_box[int(p[0])][2]:
                    pre_box[int(p[0])][0] = int(float(p[1]) * y)
                    pre_box[int(p[0])][1] = int(float(p[2]) * x)
                    pre_box[int(p[0])][2] = float(p[5])

            # calculate angle
            angles = []
            for i in range(4):
                angle1 = np.arctan((pre_box[3*i+1][0]-pre_box[3*i+2][0])/(pre_box[3*i+2][1]-pre_box[3*i+1][1]))
                angle2 = np.arctan((pre_box[3*i+0][0]-pre_box[3*i+1][0])/(pre_box[3*i+1][1]-pre_box[3*i+0][1]))
                angles.append(angle2)
                angles.append((angle1+angle2)/2)
                angles.append(angle1)
            print(angles)
        #     for i in range(12):
        #         if pre_box[i][2] == 0:
        #             continue
        #         else:
        #             distance = abs(truth_box[i]-angles[i])
        #             ws.cell(row=pointer, column=i + 13).value = distance
#
# wb.save('D:/高源/BU/AICV Lab/Hand_OA/angle_error3.xlsx')
'''


# find the fingertip from mask
'''
mask_dir = 'C:/代码/BU/hand_seg/ZiLong/data/model3/test/result'
save_dir = 'C:\代码\BU\hand_seg\YuanGao\yolov5-master\My data/fingertips_detect2'
scan_interval = 1
window_size = 7

num_of_fingertips = []
for filename in os.listdir(mask_dir):
    if filename == '9030418_pred.png':
        break
    print(filename)

    # filename = '9008322_pred.png'

    mask = cv2.imdecode(np.fromfile(mask_dir + '/' + filename, dtype=np.uint8), 1)

    # cut+padding+resize
    a = np.where(mask == 255)
    x1, x2, y1, y2 = min(a[0]), max(a[0]), min(a[1]), max(a[1])
    mask2 = mask[x1:x2, y1:y2]
    mask2 = cv2.copyMakeBorder(mask2, 50, 50, 50, 50, cv2.BORDER_CONSTANT, value=(0, 0, 0))
    dim = (200, 300)
    mask = cv2.resize(mask2, dim, interpolation=cv2.INTER_AREA)
    (y, x) = dim

    # open,close operation
    # kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (5, 5))
    # opened1 = cv2.morphologyEx(mask, cv2.MORPH_OPEN, kernel, iterations=1)     #开运算,先腐蚀后膨胀
    # closed1 = cv2.morphologyEx(mask, cv2.MORPH_CLOSE, kernel, iterations=10)    #闭运算，先膨胀后腐蚀
    # mask = closed1

    # detect edges
    fingertips = []
    edge = []
    for j in range(0, y, scan_interval):
        if 255 in mask[:, j]:
            i = np.where(mask[:, j] == 255)
            edge.append(min(i[0]))
        else:
            edge.append(300)

    for i in range(window_size, len(edge)-window_size):
        if edge[i] == 300:
            continue
        left = np.mean(edge[i-window_size:i])
        right = np.mean(edge[i:i+window_size])
        if left >= edge[i] and right >= edge[i]:
            fingertips.append(i)
    # print(fingertips)

    # detect 5 fingertips
    i = 0
    j = 1
    points = []
    while j < len(fingertips):
        while fingertips[j] < fingertips[j-1] + 10:
            j += 1
            if j == len(fingertips):
                break

        point = int((fingertips[j-1]+fingertips[i])/2)
        points.append(point)
        mask[edge[point]:edge[point]+3, point:point+3, 1] = 0

        if j == len(fingertips)-1:
            point = fingertips[j]
            points.append(point)
            mask[edge[point]:edge[point] + 3, point:point + 3, 1] = 0
            break

        i = j
        j += 1

    # print(points)

    num_of_fingertips.append(len(points))
    if len(points) < 5:
        print(filename, 4)
        plt.imshow(mask)
        plt.title(str(len(points)), fontsize='xx-large')
        plt.savefig(save_dir + '/' + filename + '.png')
    if len(points) > 5:
        print(filename, 6)
        plt.imshow(mask)
        plt.title(str(len(points)), fontsize='xx-large')
        plt.savefig(save_dir + '/' + filename + '.png')

plt.clf()
plt.hist(num_of_fingertips)
plt.show()

# plt.subplot(1,2,1)
# plt.imshow(mask, 'gray')
# plt.subplot(1,2,2)
# plt.imshow(mask)
# plt.show()

'''


# Calculate mAP
# prediction: joint detection + angle estimate (connect joint center but not fingertips)
'''
def calculate_iou(image1, image2):
    # Assume that images are binary: 255 for object, 0 for background
    # And input images are of the same size
    image1 = image1 / 255
    image2 = image2 / 255

    intersection = np.logical_and(image1, image2)
    union = np.logical_or(image1, image2)
    iou_score = np.sum(intersection) / np.sum(union)
    return iou_score

def create_image_list(txt_address):
    imgs_pred = []
    with open(txt_address, "r") as f2:
        prediction = f2.readlines()
        pre_box = np.zeros((12, 3))
        angle_box = np.zeros((4, 2))

        # extract txt to pre_box
        for p in prediction:
            p = p.split()
            if len(p) > 5:
                if float(p[5]) > pre_box[int(p[0])][2]:
                    pre_box[int(p[0])][0] = int(float(p[1]) * y)
                    pre_box[int(p[0])][1] = int(float(p[2]) * x)
                    pre_box[int(p[0])][2] = float(p[5])
            else:
                pre_box[int(p[0])][0] = int(float(p[1]) * y)
                pre_box[int(p[0])][1] = int(float(p[2]) * x)
        # calculate angle
        for i in range(4):
            angle1 = np.arctan((pre_box[3 * i + 1][0] - pre_box[3 * i + 2][0]) / (
                        pre_box[3 * i + 2][1] - pre_box[3 * i + 1][1])) / np.pi * 180
            angle2 = np.arctan((pre_box[3 * i + 0][0] - pre_box[3 * i + 1][0]) / (
                        pre_box[3 * i + 1][1] - pre_box[3 * i + 0][1])) / np.pi * 180
            angle_box[i][0] = angle1
            angle_box[i][1] = angle2
        # draw box
        for i in range(4):
            img0 = np.zeros((x, y))
            rect = ((pre_box[3 * i + 0][0], pre_box[3 * i + 0][1]), box_size, angle_box[i][1])
            box = cv2.boxPoints(rect)
            box = np.int0(box)
            imgs_pred.append(cv2.drawContours(img0, [box], 0, (255, 255, 255), -1))

            img0 = np.zeros((x, y))
            rect = ((pre_box[3 * i + 1][0], pre_box[3 * i + 1][1]), box_size, (angle_box[i][0] + angle_box[i][1]) / 2)
            box = cv2.boxPoints(rect)
            box = np.int0(box)
            imgs_pred.append(cv2.drawContours(img0, [box], 0, (255, 255, 255), -1))

            img0 = np.zeros((x, y))
            rect = ((pre_box[3 * i + 2][0], pre_box[3 * i + 2][1]), box_size, angle_box[i][0])
            box = cv2.boxPoints(rect)
            box = np.int0(box)
            imgs_pred.append(cv2.drawContours(img0, [box], 0, (255, 255, 255), -1))

    return imgs_pred

prediction_dir = 'C:/代码/BU/hand_seg/YuanGao/yolov5-master/runs/detect/test3/labels'
ground_truth_dir = 'D:/高源/BU/AICV Lab/Hand_OA/new_label'
img_dir = 'D:/高源/BU/AICV Lab/Hand_OA/3557 image+label'
box_size = (150, 150)
iou_thresh = 0.5

# excel load
wb = openpyxl.Workbook()
ws = wb.active
ws.column_dimensions['A'].width = 20
title = ['mcp2','pip2','dip2','mcp3','pip3','dip3','mcp4','pip4','dip4','mcp5','pip5','dip5']
ws.cell(row=2, column=1).value = 'mAP'
for i in range(len(title)):
    ws.cell(row=1, column=i+2).value = title[i]

# loop
row_pointer = 2
num_files = len(os.listdir(prediction_dir))
mAP_list = []
iou_list = []
for i in range(10):
    mAP_list.append(np.zeros((num_files, 12)))

for filename in os.listdir(prediction_dir):
    row_pointer += 1
    filename = filename[:-4]
    # filename = '9022408_v06'
    # if filename == '9002316_v06':
    #     break

    # write name in Excel
    ws.cell(row=row_pointer, column=1).value = filename

    # read img size
    img = cv2.imdecode(np.fromfile(img_dir + '/' + filename + '.png', dtype=np.uint8), -1)
    (x, y) = img.shape

    # create 12*2 images for each sample
    pred_txt_address = prediction_dir + '/' + filename + '.txt'
    gt_txt_address = ground_truth_dir + '/' + filename + '.txt'
    imgs_pred = create_image_list(pred_txt_address)
    imgs_gt = create_image_list(gt_txt_address)

    # calculate IOU
    for i in range(12):
        iou = calculate_iou(imgs_pred[i], imgs_gt[i])
        iou_list.append(iou)
        if iou > iou_thresh:
            ws.cell(row=row_pointer, column=i + 2).value = 1
        else:
            ws.cell(row=row_pointer, column=i + 2).value = 0

        for j in range(10):
            if iou > iou_thresh + 0.05 * j:
                mAP_list[j][row_pointer - 3][i] = 1
            else:
                mAP_list[j][row_pointer - 3][i] = 0

# calculate mAP for each category
np.set_printoptions(precision=3)
for i in range(12):
    ws.cell(row=2, column=i+2).value = np.mean(mAP_list[0], axis=0)[i]
pointer = 45
mAP_050_to_095 = np.zeros(12)
for j in range(10):
    pointer += 5
    mAP = mAP_list[j]
    ave = np.mean(mAP, axis=0)
    mAP_050_to_095 = mAP_050_to_095 + ave
    print('mAP_0'+str(pointer)+':', ave)
print('mAP_050_to_095:', mAP_050_to_095/10)
print('iou_mean:', np.mean(iou_list))

# save mAP_050 as excel
wb.save('D:/高源/BU/AICV Lab/Hand_OA/mAP_050_test.xlsx')
'''


# Draw Histogram for angle error and distance

# Distance
# distances = []
# wb = openpyxl.load_workbook('D:/高源/BU/AICV Lab/Hand_OA/Voxel_size3.xlsx')
# ws = wb.active
#
# for r in range(2, ws.max_row-1):
#     for c in range(13, 25):
#         if ws.cell(row=r, column=c).value == None:
#             continue
#         else:
#             distances.append(float(ws.cell(row=r, column=c).value))
#
# n, bins, patches = plt.hist(distances, bins=11, rwidth=0.8, align='left')
# for i in range(len(n)):
#     plt.text(bins[i], n[i]*1.02, int(n[i]), fontsize=20, horizontalalignment="center")
# plt.title('Euclidean Distance', fontsize=30)
# plt.show()

# Angle
angles = []
test_result_dir = 'C:/代码/BU/hand_seg/YuanGao/yolov5-master/runs/detect/test3/labels'
wb = openpyxl.load_workbook('D:/高源/BU/AICV Lab/Hand_OA/angle_error3.xlsx')
ws = wb.active

r = 2
for filename in os.listdir(test_result_dir):
    filename = filename[:-8]
    while r < ws.max_row:
        if str(ws.cell(row=r, column=1).value) == filename:
            for c in range(13, 25):
                if ws.cell(row=r, column=c).value == None:
                    continue
                else:
                    angles.append(float(ws.cell(row=r, column=c).value))
            r += 1
            break
        else:
            r += 1

n, bins, patches = plt.hist(angles, bins=11, rwidth=0.8, align='left')
for i in range(len(n)):
    plt.text(bins[i], n[i]*1.02, int(n[i]), fontsize=20, horizontalalignment="center")
plt.title('Angle Bias', fontsize=30)
plt.show()
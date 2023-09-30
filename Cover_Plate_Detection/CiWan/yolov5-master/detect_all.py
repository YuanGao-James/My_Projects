import argparse
import os
import sys
from pathlib import Path
from models.common import DetectMultiBackend
from utils.datasets import IMG_FORMATS, VID_FORMATS, LoadImages, LoadStreams
from utils.general import (LOGGER, check_file, check_img_size, check_imshow, check_requirements, colorstr, cv2,
                           increment_path, non_max_suppression, print_args, scale_coords, strip_optimizer, xyxy2xywh)
from utils.plots import Annotator, colors, save_one_box
from utils.torch_utils import select_device, time_sync
from utils.utils import (cvtColor, letterbox_image,preprocess_input)
import openpyxl
import numpy as np
import torch
from PIL import Image
from nets import get_model_from_name


# 路径设置
FILE = Path(__file__).resolve()
ROOT = FILE.parents[0]  # YOLOv5 root directory
if str(ROOT) not in sys.path:
    sys.path.append(str(ROOT))  # add ROOT to PATH
ROOT = Path(os.path.relpath(ROOT, Path.cwd()))  # relative


# detect函数
@torch.no_grad()
def detect(
        weights=ROOT / 'runs/train/exp11/weights/last.pt',  # model.pt path(s)
        source=ROOT / 'dir',  # file/dir/URL/glob, 0 for webcam
        data=ROOT / 'data/CW.yaml',  # dataset.yaml path
        conf_thres=0.01,  # confidence threshold
        iou_thres=0.45,  # NMS IOU threshold
        project=ROOT / 'runs/detect',  # save results to project/name
        name='exp',  # save results to project/name
        device='',  # cuda device, i.e. 0 or 0,1,2,3 or cpu
        imgsz=(640, 640),  # inference size (height, width)
        max_det=1000,  # maximum detections per image
        view_img=True,  # show results
        save_txt=True,  # save results to *.txt
        save_conf=False,  # save confidences in --save-txt labels
        save_crop=False,  # save cropped prediction boxes
        nosave=False,  # do not save images/videos
        classes=None,  # filter by class: --class 0, or --class 0 2 3
        agnostic_nms=False,  # class-agnostic NMS
        augment=False,  # augmented inference
        visualize=False,  # visualize features
        update=False,  # update all models
        line_thickness=3,  # bou
        dnn=False,  # use OpenCV DNN for ONNX inferencending box thickness (pixels)
        hide_labels=False,  # hide labels
        hide_conf=False,  # hide confidences
        half=False,  # use FP16 half-precision inference
):
    source = str(source)
    save_img = not nosave and not source.endswith('.txt')  # save inference images

    save_dir = Path(project) / name
    (save_dir / 'labels' if save_txt else save_dir).mkdir(parents=True, exist_ok=True)  # make dir

    # Load model
    device = select_device(device)
    model = DetectMultiBackend(weights, device=device, dnn=dnn, data=data, fp16=half)
    stride, names, pt = model.stride, model.names, model.pt
    imgsz = check_img_size(imgsz, s=stride)  # check image size

    # Dataloader
    dataset = LoadImages(source, img_size=imgsz, stride=stride, auto=pt)
    bs = 1  # batch_size

    # Run inference
    model.warmup(imgsz=(1 if pt else bs, 3, *imgsz))  # warmup
    dt, seen = [0.0, 0.0, 0.0], 0
    for path, im, im0s, vid_cap, s in dataset:
        t1 = time_sync()
        im = torch.from_numpy(im).to(device)
        im = im.half() if model.fp16 else im.float()  # uint8 to fp16/32
        im /= 255  # 0 - 255 to 0.0 - 1.0
        if len(im.shape) == 3:
            im = im[None]  # expand for batch dim
        t2 = time_sync()
        dt[0] += t2 - t1

        # Inference
        visualize = increment_path(save_dir / Path(path).stem, mkdir=True) if visualize else False
        pred = model(im, augment=augment, visualize=visualize)
        t3 = time_sync()
        dt[1] += t3 - t2

        # NMS
        pred = non_max_suppression(pred, conf_thres, iou_thres, classes, agnostic_nms, max_det=max_det)
        dt[2] += time_sync() - t3

        # Process predictions
        for i, det in enumerate(pred):  # per image
            seen += 1
            p, im0, frame = path, im0s.copy(), getattr(dataset, 'frame', 0)

            p = Path(p)  # to Path
            save_path = str(save_dir / p.name)  # im.jpg
            txt_path = str(save_dir / 'labels' / p.stem) + ('' if dataset.mode == 'image' else f'_{frame}')  # im.txt
            s += '%gx%g ' % im.shape[2:]  # print string
            gn = torch.tensor(im0.shape)[[1, 0, 1, 0]]  # normalization gain whwh
            imc = im0.copy() if save_crop else im0  # for save_crop
            annotator = Annotator(im0, line_width=line_thickness, example=str(names))
            if len(det):
                # Rescale boxes from img_size to im0 size
                det[:, :4] = scale_coords(im.shape[2:], det[:, :4], im0.shape).round()

                # Print results
                for c in det[:, -1].unique():
                    n = (det[:, -1] == c).sum()  # detections per class
                    s += f"{n} {names[int(c)]}{'s' * (n > 1)}, "  # add to string

                # Write results
                for *xyxy, conf, cls in reversed(det):
                    if save_txt:  # Write to file
                        xywh = (xyxy2xywh(torch.tensor(xyxy).view(1, 4)) / gn).view(-1).tolist()  # normalized xywh
                        line = (cls, *xywh, conf) if save_conf else (cls, *xywh)  # label format
                        with open(txt_path + '.txt', 'a') as f:
                            f.write(('%g ' * len(line)).rstrip() % line + '/n')

                    if save_img or save_crop or view_img:  # Add bbox to image
                        c = int(cls)  # integer class
                        label = None if hide_labels else (names[c] if hide_conf else f'{names[c]} {conf:.2f}')
                        annotator.box_label(xyxy, label, color=colors(c, True))
                        if save_crop:
                            save_one_box(xyxy, imc, file=save_dir / 'crops' / names[c] / f'{p.stem}.jpg', BGR=True)

            # Stream results
            im0 = annotator.result()

            # Save results (image with detections)
            if save_img:
                cv2.imwrite(save_path, im0)

        # Print time (inference-only)
        LOGGER.info(f'{s}Done. ({t3 - t2:.3f}s)')

    # Print results
    t = tuple(x / seen * 1E3 for x in dt)  # speeds per image
    LOGGER.info(f'Speed: %.1fms pre-process, %.1fms inference, %.1fms NMS per image at shape {(1, 3, *imgsz)}' % t)
    if save_txt or save_img:
        s = f"/n{len(list(save_dir.glob('labels/*.txt')))} labels saved to {save_dir / 'labels'}" if save_txt else ''
        LOGGER.info(f"Results saved to {colorstr('bold', save_dir)}{s}")
    if update:
        strip_optimizer(weights)  # update model (to fix SourceChangeWarning)

# 启用yolo模型进行预测
def detect_yolo():
    # detect(
    #     weights=ROOT / 'weights/CW_best2.pt',  # model.pt path(s)
    #     source=ROOT / 'Deployment_test_0419/hougai/ciwan',  # file/dir/URL/glob, 0 for webcam
    #     data=ROOT / 'data/CW.yaml',  # dataset.yaml path
    #     conf_thres=0.02,  # confidence threshold
    #     iou_thres=0.1,  # NMS IOU threshold
    #     project=ROOT / 'runs/result/hougai',  # save results to project/name
    #     name='ciwan',  # save results to project/name
    # )
    detect(
        weights=ROOT / 'weights/Front134_2.pt',  # model.pt path(s)
        source=ROOT / 'Deployment_test_0419/qiangai/front/crop/front1',  # file/dir/URL/glob, 0 for webcam
        data=ROOT / 'data/front134.yaml',  # dataset.yaml path
        conf_thres=0.2,  # confidence threshold
        iou_thres=0.1,  # NMS IOU threshold
        project=ROOT / 'runs/result/qiangai/front/crop',  # save results to project/name
        name='front1',  # save results to project/name
    )
    detect(
        weights=ROOT / 'weights/Front134_2.pt',  # model.pt path(s)
        source=ROOT / 'Deployment_test_0419/qiangai/front/crop/front3',  # file/dir/URL/glob, 0 for webcam
        data=ROOT / 'data/front134.yaml',  # dataset.yaml path
        conf_thres=0.25,  # confidence threshold
        iou_thres=0.1,  # NMS IOU threshold
        project=ROOT / 'runs/result/qiangai/front/crop',  # save results to project/name
        name='front3',  # save results to project/name
    )
    detect(
        weights=ROOT / 'weights/Front134_2.pt',  # model.pt path(s)
        source=ROOT / 'Deployment_test_0419/qiangai/front/crop/front4',  # file/dir/URL/glob, 0 for webcam
        data=ROOT / 'data/front134.yaml',  # dataset.yaml path
        conf_thres=0.2,  # confidence threshold
        iou_thres=0.1,  # NMS IOU threshold
        project=ROOT / 'runs/result/qiangai/front/crop',  # save results to project/name
        name='front4',  # save results to project/name
    )
    detect(
        weights=ROOT / 'weights/front2_best.pt',  # model.pt path(s)
        source=ROOT / 'Deployment_test_0419/qiangai/front/crop/front2',  # file/dir/URL/glob, 0 for webcam
        data=ROOT / 'data/front2.yaml',  # dataset.yaml path
        conf_thres=0.2,  # confidence threshold
        iou_thres=0.1,  # NMS IOU threshold
        project=ROOT / 'runs/result/qiangai/front/crop',  # save results to project/name
        name='front2',  # save results to project/name
    )

# 读取后盖label写入excel
def write_names_hougai(directory_name, ws):
        for filename in os.listdir(directory_name):
            k = 1
            if filename == 'labels':
                continue
            while 1:
                if ws.cell(row=k, column=1).value == None:
                    ws.cell(row=k, column=1).value = filename[:17]
                    break
                elif filename[:17] != ws.cell(row=k, column=1).value:
                    k += 1
                else:
                    break

def yolo2excel(directory_name, num, ws):
    for filename in os.listdir(directory_name+'/labels'):
        k = 1
        if filename == 'Thumbs':
            continue
        while 1:
            if filename[:17] != ws.cell(row=k, column=1).value:
                k += 1
            else:
                break
            if ws.cell(row=k, column=1).value == None:
                ws.cell(row=k, column=1).value = filename[:17]
                break
        ws.cell(row=k, column=num).value = 1

def detect_image(image, model, thre=0.5):
    input_shape = [224, 224]
    class_names = ['NG', 'OK']
    # ---------------------------------------------------------#
    #   在这里将图像转换成RGB图像，防止灰度图在预测时报错。
    #   代码仅仅支持RGB图像的预测，所有其它类型的图像都会转化成RGB
    # ---------------------------------------------------------#
    image = cvtColor(image)
    # ---------------------------------------------------#
    #   对图片进行不失真的resize
    # ---------------------------------------------------#
    image_data = letterbox_image(image, [input_shape[1], input_shape[0]], letterbox_image=True)
    # ---------------------------------------------------------#
    #   归一化+添加上batch_size维度+转置
    # ---------------------------------------------------------#
    image_data = np.transpose(np.expand_dims(preprocess_input(np.array(image_data, np.float32)), 0), (0, 3, 1, 2))

    with torch.no_grad():
        photo = torch.from_numpy(image_data)
        # ---------------------------------------------------#
        #   图片传入网络进行预测
        # ---------------------------------------------------#
        preds = torch.softmax(model(photo)[0], dim=-1).cpu().numpy()
    # ---------------------------------------------------#
    #   获得所属种类
    # ---------------------------------------------------#
    if preds[0] > thre:
        class_name = class_names[0]
    else:
        class_name = class_names[1]
    # class_name  = self.class_names[np.argmax(preds)]

    return class_name

def resnet2excel(ws):
    backbone = 'resnet50'
    for dir_name in os.listdir('C:/Users/ASUS/Desktop/Using_AI/CiWan/yolov5-master/Deployment_test_0419/hougai/Screw'):
        if dir_name == 'LeftDown' or dir_name == 'LeftUp':
            model_path = 'logs/mix.pth'
            print('stage1')
        else:
            model_path = 'logs/M_and_R.pth'
            print('stage2')
        model = get_model_from_name[backbone](num_classes=2, pretrained=False)
        device = torch.device('cpu')
        model.load_state_dict(torch.load(model_path, map_location=device))
        model = model.eval()
        for filename in os.listdir('C:/Users/ASUS/Desktop/Using_AI/CiWan/yolov5-master/Deployment_test_0419/hougai/Screw/'+dir_name):
            if dir_name == 'LeftDown':
                num = 3
            elif dir_name == 'LeftUp':
                num = 4
            elif dir_name == 'Middle':
                num = 5
            elif dir_name == 'Right':
                num = 6
            else:
                num = 15
            image = Image.open('C:/Users/ASUS/Desktop/Using_AI/CiWan/yolov5-master/Deployment_test_0419/hougai/Screw/'+dir_name
                               + '/' + filename)
            class_name = detect_image(image, model, thre=0.7)
            k = 2
            while 1:
                if ws.cell(row=k, column=1).value == None:
                    ws.cell(row=k, column=1).value = filename[:17]
                    break
                elif filename[:17] != ws.cell(row=k, column=1).value:
                    k += 1
                else:
                    break
            if class_name == 'NG':
                ws.cell(row=k, column=num).value = 1

def predict_hougai(ws):
    for i in range(ws.max_row+1):
        if i < 2:
            continue
        elif ws.cell(row=i, column=2).value == 1 or ws.cell(row=i, column=3).value == 1 or ws.cell(row=i, column=4).value == 1\
                or ws.cell(row=i, column=5).value == 1 or ws.cell(row=i, column=6).value == 1:
            ws.cell(row=i, column=7).value = 1
        else:
            pass

def hougai():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 15
    ws.column_dimensions['K'].width = 15
    ws.column_dimensions['L'].width = 15
    ws.column_dimensions['M'].width = 15
    ws.column_dimensions['N'].width = 15
    ws.column_dimensions['O'].width = 15
    title1 = ["大图名", "CW", "LeftDown", "LeftUp", "Middle", "Right", "后盖预测", '真实值', '总漏杀', '总过杀', '过杀cw', '过杀ld', '过杀lu', '过杀m', '过杀r']
    for i in range(len(title1)):
        ws.cell(row=1, column=i + 1).value = title1[i]
    write_names_hougai('C:/Users/ASUS/Desktop/Using_AI/CiWan/yolov5-master/runs/result/hougai/ciwan', ws)
    yolo2excel('C:/Users/ASUS/Desktop/Using_AI/CiWan/yolov5-master/runs/result/hougai/ciwan', 2, ws)
    resnet2excel(ws)
    predict_hougai(ws)

    # 读取真实值写入excel
    for filename in os.listdir('C:/Users/ASUS/Desktop/Using AI/瓷碗目标检测/有整体标签测试数据4.16/4.16hougai/NG/Ciwan'):
        k = 2
        while 1:
            if filename[:17] == ws.cell(row=k, column=1).value:
                ws.cell(row=k, column=8).value = 1
                break
            else:
                k += 1
    for filename in os.listdir('C:/Users/ASUS/Desktop/Using AI/瓷碗目标检测/有整体标签测试数据4.16/hougailuosikong/NG/Screw/LeftDown'):
        k = 2
        while 1:
            if filename[:17] == ws.cell(row=k, column=1).value:
                ws.cell(row=k, column=8).value = 1
                break
            else:
                k += 1
    for filename in os.listdir('C:/Users/ASUS/Desktop/Using AI/瓷碗目标检测/有整体标签测试数据4.16/hougailuosikong/NG/Screw/LeftUp'):
        k = 2
        while 1:
            if filename[:17] == ws.cell(row=k, column=1).value:
                ws.cell(row=k, column=8).value = 1
                break
            else:
                k += 1
    for filename in os.listdir('C:/Users/ASUS/Desktop/Using AI/瓷碗目标检测/有整体标签测试数据4.16/hougailuosikong/NG/Screw/Middle'):
        k = 2
        while 1:
            if filename[:17] == ws.cell(row=k, column=1).value:
                ws.cell(row=k, column=8).value = 1
                break
            else:
                k += 1
    for filename in os.listdir('C:/Users/ASUS/Desktop/Using AI/瓷碗目标检测/有整体标签测试数据4.16/hougailuosikong/NG/Screw/Right'):
        k = 2
        while 1:
            if filename[:17] == ws.cell(row=k, column=1).value:
                ws.cell(row=k, column=8).value = 1
                break
            else:
                k += 1

    # 统计过杀率，漏杀率
    lousha = 0
    guosha = 0
    cw = 0
    ld = 0
    lu = 0
    m = 0
    r = 0
    for i in range(2, ws.max_row + 1):
        if ws.cell(row=i, column=8).value == 1:
            if ws.cell(row=i, column=7).value != 1:
                lousha += 1
            else:
                pass
        else:
            if ws.cell(row=i, column=7).value == 1:
                guosha += 1
            if ws.cell(row=i, column=2).value == 1:
                cw += 1
            if ws.cell(row=i, column=3).value == 1:
                ld += 1
            if ws.cell(row=i, column=4).value == 1:
                lu += 1
            if ws.cell(row=i, column=5).value == 1:
                m += 1
            if ws.cell(row=i, column=6).value == 1:
                r += 1
    ws.cell(row=2, column=9).value = str(lousha) + '/' + str(ws.max_row - 1) + '=' + str(
        round(lousha / (ws.max_row - 1), 3))
    ws.cell(row=2, column=10).value = str(guosha) + '/' + str(ws.max_row - 1) + '=' + str(
        round(guosha / (ws.max_row - 1), 3))
    ws.cell(row=2, column=11).value = str(cw) + '/' + str(ws.max_row - 1) + '=' + str(
        round(cw / (ws.max_row - 1), 3))
    ws.cell(row=2, column=12).value = str(ld) + '/' + str(ws.max_row - 1) + '=' + str(
        round(ld / (ws.max_row - 1), 3))
    ws.cell(row=2, column=13).value = str(lu) + '/' + str(ws.max_row - 1) + '=' + str(
        round(lu / (ws.max_row - 1), 3))
    ws.cell(row=2, column=14).value = str(m) + '/' + str(ws.max_row - 1) + '=' + str(
        round(m / (ws.max_row - 1), 3))
    ws.cell(row=2, column=15).value = str(r) + '/' + str(ws.max_row - 1) + '=' + str(
        round(r / (ws.max_row - 1), 3))

    wb.save('后盖预测.xlsx')

# 读取前盖label写入excel
def write_names_qiangai(directory_name, ws):
    for filename in os.listdir(directory_name):
        k = 1
        if filename == 'labels':
            continue
        while 1:
            if ws.cell(row=k, column=1).value == None:
                ws.cell(row=k, column=1).value = filename[:18]
                break
            elif filename[:18] != ws.cell(row=k, column=1).value:
                k += 1
            else:
                break

def label2excel2(directory_name, num, ws2):
    k = 2
    for filename in os.listdir(directory_name):
        if filename == 'labels':
            continue
        if filename[:18]+'.txt' in os.listdir(directory_name+'/labels'):
            ws2.cell(row=k, column=num).value = 1
        else:
            pass
        k += 1

def predict_qiangai(ws):
    for i in range(ws.max_row + 1):
        if i < 2:
            continue
        elif ws.cell(row=i, column=2).value == 1 or ws.cell(row=i, column=3).value == 1 or ws.cell(row=i, column \
                =4).value == 1 or ws.cell(row=i, column=5).value == 1:
            ws.cell(row=i, column=6).value = 1
        else:
            pass

def qiangai():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 15
    ws.column_dimensions['K'].width = 15
    ws.column_dimensions['L'].width = 15
    ws.column_dimensions['M'].width = 15
    title2 = ["大图名", "front1", "front2", "front3", "front4", "前盖预测", '真实值', '总漏杀', '总过杀', '过杀1', '过杀2', '过杀3', '过杀4']
    for i in range(len(title2)):
        ws.cell(row=1, column=i + 1).value = title2[i]
    write_names_qiangai('C:/Users/ASUS/Desktop/Using AI/瓷碗目标检测/前盖大图/4.16qiangaiNG/Front/Origin/Light2', ws)
    write_names_qiangai('C:/Users/ASUS/Desktop/Using AI/瓷碗目标检测/前盖大图/4.16qiangaiOK/Front/Origin/Light2', ws)
    ws.auto_filter.add_sort_condition('A2:A407')
    label2excel2('C:/Users/ASUS/Desktop/Using_AI/CiWan/yolov5-master/runs/result/qiangai/front/crop/front1', 2, ws)
    label2excel2('C:/Users/ASUS/Desktop/Using_AI/CiWan/yolov5-master/runs/result/qiangai/front/crop/front2', 3, ws)
    label2excel2('C:/Users/ASUS/Desktop/Using_AI/CiWan/yolov5-master/runs/result/qiangai/front/crop/front3', 4, ws)
    label2excel2('C:/Users/ASUS/Desktop/Using_AI/CiWan/yolov5-master/runs/result/qiangai/front/crop/front4', 5, ws)
    predict_qiangai(ws)

    # 读取真实值写入excel
    for filename in os.listdir('C:/Users/ASUS/Desktop/Using AI/瓷碗目标检测/前盖大图/4.16qiangaiNG/Front/Origin/Light2'):
        k = 2
        while 1:
            if filename[:18] == ws.cell(row=k, column=1).value:
                ws.cell(row=k, column=7).value = 1
                break
            else:
                k += 1
    # 统计过杀率，漏杀率
    lousha = 0
    guosha = 0
    guosha1 = 0
    guosha2 = 0
    guosha3 = 0
    guosha4 = 0
    for i in range(2, ws.max_row + 1):
        if ws.cell(row=i, column=7).value == 1:
            if ws.cell(row=i, column=6).value != 1:
                lousha += 1
            else:
                pass
        else:
            if ws.cell(row=i, column=6).value == 1:
                guosha += 1
            if ws.cell(row=i, column=2).value == 1:
                guosha1 += 1
            if ws.cell(row=i, column=3).value == 1:
                guosha2 += 1
            if ws.cell(row=i, column=4).value == 1:
                guosha3 += 1
            if ws.cell(row=i, column=5).value == 1:
                guosha4 += 1
    ws.cell(row=2, column=8).value = str(lousha) + '/' + str(ws.max_row-1) + '=' + str(round(lousha/(ws.max_row-1), 3))
    ws.cell(row=2, column=9).value = str(guosha) + '/' + str(ws.max_row-1) + '=' + str(round(guosha/(ws.max_row-1), 3))
    ws.cell(row=2, column=10).value = str(guosha1) + '/' + str(ws.max_row-1) + '=' + str(round(guosha1/(ws.max_row-1), 3))
    ws.cell(row=2, column=11).value = str(guosha2) + '/' + str(ws.max_row-1) + '=' + str(round(guosha2/(ws.max_row-1), 3))
    ws.cell(row=2, column=12).value = str(guosha3) + '/' + str(ws.max_row-1) + '=' + str(round(guosha3/(ws.max_row-1), 3))
    ws.cell(row=2, column=13).value = str(guosha4) + '/' + str(ws.max_row-1) + '=' + str(round(guosha4/(ws.max_row-1), 3))

    wb.save('前盖预测.xlsx')

if __name__ == '__main__':
    detect_yolo()
    hougai()
    qiangai()

import argparse
import os
import random
import sys
from pathlib import Path
import matplotlib
import matplotlib.pyplot as plt

from models.common import DetectMultiBackend
from utils.datasets import IMG_FORMATS, VID_FORMATS, LoadImages, LoadStreams
from utils.general import (LOGGER, check_file, check_img_size, check_imshow, check_requirements, colorstr, cv2,
                           increment_path, non_max_suppression, print_args, scale_coords, strip_optimizer, xyxy2xywh)
from utils.plots import Annotator, colors, save_one_box
from utils.torch_utils import select_device, time_sync
from utils.utils import (cvtColor, letterbox_image,preprocess_input)
import openpyxl
from openpyxl.styles import Alignment
import numpy as np
import torch
from PIL import Image
from feature_matching_and_xml2txt import matching_output_result, xml_convert
from plot_box_on_pic import plot_box_on_pic
import xml.etree.ElementTree as ET

# 路径设置
FILE = Path(__file__).resolve()
ROOT = FILE.parents[0]  # YOLOv5 root directory
if str(ROOT) not in sys.path:
    sys.path.append(str(ROOT))  # add ROOT to PATH
ROOT = Path(os.path.relpath(ROOT, Path.cwd()))  # relative

class_names = ['bian', 'heidian', 'liaohua', 'guankouquejiao', 'duanzhen', 'wuneiguan', 'shuikougao', 'neiguanquejiao',
  'baiseqipao', 'shuiyin', 'tuoshang', 'jiaxian', 'yueyayashang', 'youlieheng', 'guankoupifeng', 'toubuchuankong',
  'toubupifeng', 'toubusuoshui', 'others', 'background']

# detect函数
@torch.no_grad()
def detect(
        weights=ROOT / 'runs/train/exp11/weights/last.pt',  # model.pt path(s)
        source=ROOT / 'dir',  # file/dir/URL/glob, 0 for webcam
        data=ROOT / 'data/CW.yaml',  # dataset.yaml path
        conf_thres=0.01,  # confidence threshold
        iou_thres=0.45,  # NMS IOU threshold
        project=ROOT / 'runs/detect',  # save results to project/name
        name='exp',  # save results to project/name
        device='',  # cuda device, i.e. 0 or 0,1,2,3 or cpu
        imgsz=(640, 640),  # inference size (height, width)
        max_det=1000,  # maximum detections per image
        view_img=True,  # show results
        save_txt=True,  # save results to *.txt
        save_conf=False,  # save confidences in --save-txt labels
        save_crop=False,  # save cropped prediction boxes
        nosave=False,  # do not save images/videos
        classes=None,  # filter by class: --class 0, or --class 0 2 3
        agnostic_nms=False,  # class-agnostic NMS
        augment=False,  # augmented inference
        visualize=False,  # visualize features
        update=False,  # update all models
        line_thickness=3,  # bou
        dnn=False,  # use OpenCV DNN for ONNX inferencending box thickness (pixels)
        hide_labels=False,  # hide labels
        hide_conf=False,  # hide confidences
        half=False,  # use FP16 half-precision inference
):
    source = str(source)
    save_img = not nosave and not source.endswith('.txt')  # save inference images

    save_dir = Path(project) / name
    (save_dir / 'labels' if save_txt else save_dir).mkdir(parents=True, exist_ok=True)  # make dir

    # Load model
    device = select_device(device)
    model = DetectMultiBackend(weights, device=device, dnn=dnn, data=data, fp16=half)
    stride, names, pt = model.stride, model.names, model.pt
    imgsz = check_img_size(imgsz, s=stride)  # check image size

    # Dataloader
    dataset = LoadImages(source, img_size=imgsz, stride=stride, auto=pt)
    bs = 1  # batch_size

    # Run inference
    model.warmup(imgsz=(1 if pt else bs, 3, *imgsz))  # warmup
    dt, seen = [0.0, 0.0, 0.0], 0
    for path, im, im0s, vid_cap, s in dataset:
        t1 = time_sync()
        im = torch.from_numpy(im).to(device)
        im = im.half() if model.fp16 else im.float()  # uint8 to fp16/32
        im /= 255  # 0 - 255 to 0.0 - 1.0
        if len(im.shape) == 3:
            im = im[None]  # expand for batch dim
        t2 = time_sync()
        dt[0] += t2 - t1

        # Inference
        visualize = increment_path(save_dir / Path(path).stem, mkdir=True) if visualize else False
        pred = model(im, augment=augment, visualize=visualize)
        t3 = time_sync()
        dt[1] += t3 - t2

        # NMS
        pred = non_max_suppression(pred, conf_thres, iou_thres, classes, agnostic_nms, max_det=max_det)
        dt[2] += time_sync() - t3

        # Process predictions
        for i, det in enumerate(pred):  # per image
            seen += 1
            p, im0, frame = path, im0s.copy(), getattr(dataset, 'frame', 0)

            p = Path(p)  # to Path
            save_path = str(save_dir / p.name)  # im.jpg
            txt_path = str(save_dir / 'labels' / p.stem) + ('' if dataset.mode == 'image' else f'_{frame}')  # im.txt
            s += '%gx%g ' % im.shape[2:]  # print string
            gn = torch.tensor(im0.shape)[[1, 0, 1, 0]]  # normalization gain whwh
            imc = im0.copy() if save_crop else im0  # for save_crop
            annotator = Annotator(im0, line_width=line_thickness, example=str(names))
            if len(det):
                # Rescale boxes from img_size to im0 size
                det[:, :4] = scale_coords(im.shape[2:], det[:, :4], im0.shape).round()

                # Print results
                for c in det[:, -1].unique():
                    n = (det[:, -1] == c).sum()  # detections per class
                    s += f"{n} {names[int(c)]}{'s' * (n > 1)}, "  # add to string

                # Write results
                for *xyxy, conf, cls in reversed(det):
                    if int(cls) == 4:
                        if int(xyxy[1]) < 180 or int(xyxy[3]) > 250:
                            continue
                    if int(cls) == 1:
                        if conf < 0.3:
                            continue
                    if save_txt:  # Write to file
                        xywh = (xyxy2xywh(torch.tensor(xyxy).view(1, 4)) / gn).view(-1).tolist()  # normalized xywh

                        line = (cls, *xywh, conf) if save_conf else (cls, *xywh)  # label format
                        with open(txt_path + '.txt', 'a') as f:
                            f.write(('%g ' * len(line)).rstrip() % line + '\n')

                    if save_img or save_crop or view_img:  # Add bbox to image
                        c = int(cls)  # integer class
                        label = None if hide_labels else (names[c] if hide_conf else f'{names[c]} {conf:.2f}')
                        annotator.box_label(xyxy, label, color=colors(c, True))
                        if save_crop:
                            save_one_box(xyxy, imc, file=save_dir / 'crops' / names[c] / f'{p.stem}.jpg', BGR=True)

            # Stream results
            im0 = annotator.result()

            # Save results (image with detections)
            if save_img:
                cv2.imwrite(save_path, im0)

        # Print time (inference-only)
        LOGGER.info(f'{s}Done. ({t3 - t2:.3f}s)')

    # Print results
    t = tuple(x / seen * 1E3 for x in dt)  # speeds per image
    LOGGER.info(f'Speed: %.1fms pre-process, %.1fms inference, %.1fms NMS per image at shape {(1, 3, *imgsz)}' % t)
    if save_txt or save_img:
        s = f"\n{len(list(save_dir.glob('labels/*.txt')))} labels saved to {save_dir / 'labels'}" if save_txt else ''
        LOGGER.info(f"Results saved to {colorstr('bold', save_dir)}{s}")
    if update:
        strip_optimizer(weights)  # update model (to fix SourceChangeWarning)

# 启用yolo模型进行预测
def detect_YD(source,project,name,conf,weights):
    detect(
        weights=ROOT / weights,  # model.pt path(s)
        source=ROOT / source,  # file/dir/URL/glob, 0 for webcam
        data=ROOT / 'data/YD.yaml',  # dataset.yaml path
        conf_thres=conf,  # confidence threshold
        iou_thres=0.1,  # NMS IOU threshold
        project=ROOT / project,  # save results to project/name
        name=name,  # save results to project/name
    )

# 读取label写入excel
def write_pic_names(directory_name, ws, num):
    l = os.listdir(directory_name)
    l.remove('labels')
    l.sort(key=lambda x: int(x.split('.')[0].split('_')[-1]))
    for filename in l:
        k = 2
        while 1:
            if ws.cell(row=k, column=num).value == None:
                ws.cell(row=k, column=num).value = filename[:-4]
                break
            else:
                k += 1

def write_pic_path(directory_name, ws, num):
    l = os.listdir(directory_name)
    l.remove('labels')
    l.sort(key=lambda x: int(x.split('.')[0].split('_')[-1]))
    k = 2
    for filename in l:
        ws.cell(row=k, column=num).value = directory_name+'/'+filename
        k += 1

def write_truth_by_label(directory_name, ws, num):
    l = os.listdir(directory_name)
    l.sort(key=lambda x: int(x.split('.')[0].split('_')[-1]))
    for filename in l:
        if filename[-4:] == '.bmp':
            continue
        stat_info = os.stat(directory_name + '/' + filename)
        if stat_info.st_size == 0:
            continue
        if filename[-4:] == '.txt':
            filename = filename[:-4]
        k = 2
        while 1:
            if ws.cell(row=k, column=1).value == filename or ws.cell(row=k, column=2).value == filename:
                ws.cell(row=k, column=num).value = 1
                break
            else:
                k += 1
            if ws.cell(row=k, column=1).value == None:
                print('no file:'+filename)
                break

def write_truth_by_dir(directory_name, ws, num):
    l = os.listdir(directory_name)
    for filename in l:
        if filename[-4:] == '.bmp':
            continue
        k = 2
        while 1:
            if ws.cell(row=k, column=1).value == filename:
                ws.cell(row=k, column=num).value = 1
                break
            else:
                k += 1
            if ws.cell(row=k, column=1).value == None:
                print('no file:'+filename)
                break

def write_excel_by_order(directory_name, ws, num):
    l = os.listdir(directory_name)
    l.remove('labels')
    l.sort(key=lambda x: int(x.split('.')[0].split('_')[-1]))
    k = 1
    height, width = 228, 865
    for filename in l:
        if filename != 'labels':
            k += 1
            if filename[:-4]+'.txt' in os.listdir(directory_name+'/labels'):
                ws.cell(row=k, column=9).value = 1
                with open(directory_name+'/labels/'+filename[:-4]+'.txt') as f:
                    annotations = f.readlines()
                    classes = []
                    points = []
                    for ann in annotations:
                        ann = list(map(float, ann.split()))
                        ann[0] = int(ann[0])
                        cls, x, y, w, h = ann
                        c1, c2 = (int((x - w / 2) * width), int((y - h / 2) * height)), (int((x + w / 2) * width), int((y + h / 2) * height))
                        points.append([c1, c2])
                        classes.append(class_names[ann[0]])
                p = ''
                c = ''
                for i in range(len(points)):
                    p = p + '\n' + str(points[i])
                    c = c + '\n' + str(classes[i])
                ws.cell(row=k, column=num).value = c
                ws.cell(row=k, column=num+1).value = p

def save_to_excel(dir_001,dir_210,txt_dir_001,txt_dir_210,output_path,label=None):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 30
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 30
    ws.column_dimensions['K'].width = 20
    ws.column_dimensions['L'].width = 20
    title = ['ID_001', 'ID_210', 'path_001', 'path_210', '001', '坐标', "210", '坐标', '预测值', '真实值', '漏杀', '过杀']
    for i in range(len(title)):
        ws.cell(row=1, column=i + 1).value = title[i]
    write_pic_names(dir_001, ws, 1)
    write_pic_names(dir_210, ws, 2)
    write_excel_by_order(dir_001, ws=ws, num=5)
    write_pic_path(dir_001, ws=ws, num=3)
    write_excel_by_order(dir_210, ws=ws, num=7)
    write_pic_path(dir_210, ws=ws, num=4)
    if label == None:
        write_truth_by_label(txt_dir_001, ws=ws, num=10)
        write_truth_by_label(txt_dir_210, ws=ws, num=10)
    if label == 'NG':
        for i in range(1, ws.max_row):
            ws.cell(row=i+1, column=10).value = 1
    wb.save(output_path+'.xlsx')

def mix_excel(date):
    new_wb = openpyxl.Workbook()
    new_ws = new_wb.active
    new_ws.column_dimensions['A'].width = 30
    new_ws.column_dimensions['B'].width = 30
    new_ws.column_dimensions['C'].width = 15
    new_ws.column_dimensions['D'].width = 15
    new_ws.column_dimensions['E'].width = 15
    new_ws.column_dimensions['F'].width = 30
    new_ws.column_dimensions['G'].width = 15
    new_ws.column_dimensions['H'].width = 30
    new_ws.column_dimensions['K'].width = 20
    new_ws.column_dimensions['L'].width = 20
    new_ws.column_dimensions['M'].width = 20
    title = ['ID_001', 'ID_210', 'path_001', 'path_210', '001', '坐标', "210", '坐标', '预测值', '真实值', '漏杀', '过杀', '准确率']
    for i in range(len(title)):
        new_ws.cell(row=1, column=i + 1).value = title[i]
    guosha = 0
    lousha = 0
    if 'OK' in os.listdir('D:/Using AI/烟弹管缺陷检测/Data/'+date):
        for excel_num in os.listdir('D:/Using AI/烟弹管缺陷检测/Data/'+date+'/OK'):
            if excel_num[-4:] != 'xlsx':
                continue
            else:
                wb = openpyxl.load_workbook('D:/Using AI/烟弹管缺陷检测/Data/'+date+'/OK/'+excel_num)
                ws = wb.active
                for rows in ws.iter_rows(min_row=2, values_only=True):
                    rows = list(rows)
                    if rows[8] == 1 and rows[9] == None:
                        guosha += 1
                    if rows[9] == 1 and rows[8] == None:
                        lousha += 1
                    new_ws.append(rows)
    elif 'NG' in os.listdir('D:/Using AI/烟弹管缺陷检测/Data/' + date):
        for excel_num in os.listdir('D:/Using AI/烟弹管缺陷检测/Data/'+date+'/NG'):
            if excel_num[-4:] != 'xlsx':
                continue
            else:
                wb = openpyxl.load_workbook('D:/Using AI/烟弹管缺陷检测/Data/'+date+'/NG/'+excel_num)
                ws = wb.active
                for rows in ws.iter_rows(min_row=2, values_only=True):
                    rows = list(rows)
                    if rows[8] == 1 and rows[9] == None:
                        guosha += 1
                    if rows[9] == 1 and rows[8] == None:
                        lousha += 1
                    new_ws.append(rows)
    else:
        for excel_num in os.listdir('D:/Using AI/烟弹管缺陷检测/Data/'+date):
            if excel_num[-4:] != 'xlsx':
                continue
            else:
                wb = openpyxl.load_workbook('D:/Using AI/烟弹管缺陷检测/Data/'+date+'/'+excel_num)
                ws = wb.active
                for rows in ws.iter_rows(min_row=2, values_only=True):
                    rows = list(rows)
                    if rows[8] == 1 and rows[9] == None:
                        guosha += 1
                    if rows[9] == 1 and rows[8] == None:
                        lousha += 1
                    new_ws.append(rows)
    new_ws.cell(row=2, column=11).value = str(lousha) + '/' + str(new_ws.max_row-1) + '=' + str(round(lousha/(new_ws.max_row-1),3))
    new_ws.cell(row=2, column=12).value = str(guosha) + '/' + str(new_ws.max_row-1) + '=' + str(round(guosha/(new_ws.max_row-1),3))
    new_ws.cell(row=2, column=13).value = str(new_ws.max_row-1-guosha-lousha) + '/' + str(new_ws.max_row - 1) + '=' + str(round((new_ws.max_row-1-guosha-lousha) / (new_ws.max_row - 1), 3))
    # 居中
    for key in list(new_ws._cells.keys()): new_ws._cells[key].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    new_wb.save('D:/Using AI/烟弹管缺陷检测/Data/'+date+'/'+date+'_result.xlsx')

def cut_img(input_dir, output_dir):
    os.makedirs(output_dir) if os.path.exists(output_dir) == False else 0
    matching_output_result(input_dir, output_dir)

def output_wrong_img(date):
    os.makedirs('D:/Using AI/烟弹管缺陷检测/Data/' + date + '/wrong_img/guosha') if os.path.exists('D:/Using AI/烟弹管缺陷检测/Data/' + date + '/wrong_img/guosha') == False else 0
    os.makedirs('D:/Using AI/烟弹管缺陷检测/Data/' + date + '/wrong_img/lousha') if os.path.exists('D:/Using AI/烟弹管缺陷检测/Data/' + date + '/wrong_img/lousha') == False else 0
    wb = openpyxl.load_workbook('D:/Using AI/烟弹管缺陷检测/Data/' + date + '/' + date + '_result.xlsx')
    ws = wb.active
    for i in range(ws.max_row):
        # 过杀
        if ws.cell(row=i+1, column=9).value == 1 and ws.cell(row=i+1, column=10).value == None:
            try:
                filename_210 = ws.cell(row=i+1, column=2).value
                path_001 = ws.cell(row=i+1, column=3).value
                path_210 = ws.cell(row=i+1, column=4).value
                image_001 = cv2.imdecode(np.fromfile(path_001, dtype=np.uint8), -1)
                image_210 = cv2.imdecode(np.fromfile(path_210, dtype=np.uint8), -1)
                img_concat = np.concatenate((image_001, image_210), axis=0)
                cv2.imencode('.bmp', img_concat)[1].tofile('D:/Using AI/烟弹管缺陷检测/Data/' + date + '/wrong_img/guosha/' + filename_210 + '.bmp')
            except:
                pass
        # 漏杀
        if ws.cell(row=i+1, column=9).value == None and ws.cell(row=i+1, column=10).value == 1:
            try:
                filename_210 = ws.cell(row=i+1, column=2).value
                path_001 = ws.cell(row=i+1, column=3).value
                path_210 = ws.cell(row=i+1, column=4).value
                image_001 = cv2.imdecode(np.fromfile(path_001, dtype=np.uint8), -1)
                image_210 = cv2.imdecode(np.fromfile(path_210, dtype=np.uint8), -1)
                img_concat = np.concatenate((image_001, image_210), axis=0)
                cv2.imencode('.bmp', img_concat)[1].tofile('D:/Using AI/烟弹管缺陷检测/Data/' + date + '/wrong_img/lousha/' + filename_210 + '.bmp')
            except:
                pass

def error_stat(label_directory_pre, label_directory_true=None, save_path=None):
    guosha = [0] * 20
    lousha = [0] * 20
    for num in os.listdir(label_directory_pre):
        for filename in os.listdir(label_directory_pre+'/'+num):
            if filename == 'labels':
                continue
            prediction = []
            truth = []
            if filename[:-4] + '.txt' in os.listdir(label_directory_pre+'/'+num + '/labels'):
                with open(label_directory_pre+'/'+num + '/labels/' + filename[:-4] + '.txt', 'r') as f:
                    annotations = f.readlines()
                    for ann in annotations:
                        ann = list(map(float, ann.split()))
                        if int(ann[0]) != 0:
                            prediction.append(int(ann[0]))
            if label_directory_true != None:
                try:
                    if filename[:-4] in os.listdir(label_directory_true + '/' + num):
                        tree = ET.parse(label_directory_true + '/' + num + '/' + filename[:-4])
                    else:
                        tree = ET.parse(label_directory_true + '/' + num + '/' + filename[:-4] + '.xml')
                    root = tree.getroot()
                    for obj in root.iter('object'):
                        cls_id = obj.find('name').text
                        if cls_id != '0' and cls_id not in truth:
                            truth.append(int(cls_id))
                except:
                    pass
            for p in prediction:
                if p not in truth:
                    guosha[p] += 1
            if label_directory_true != None:
                for p in truth:
                    if int(p) == 19:
                        continue
                    if p not in prediction:
                        lousha[p] += 1
    plt.figure(figsize=(18, 12))
    ax1 = plt.subplot(2, 1, 1)
    ax2 = plt.subplot(2, 1, 2)
    plt.sca(ax1)
    plt.title('guosha')
    plt.barh(class_names, guosha)
    plt.sca(ax2)
    plt.title('lousha')
    plt.barh(class_names, lousha)
    # plt.show()
    myfig = plt.gcf()
    myfig.savefig(save_path, dpi=300)

if __name__ == '__main__':
    matplotlib.use('TkAgg')

    # -------------
    # 测试无标注框的数据
    # -------------

    # date = '5.27'
    # for num in os.listdir('D:/Using AI/烟弹管缺陷检测/Data/'+date+'/A3138MG000_7K01D72PAK00210'):
        # cut_img('D:/Using AI/烟弹管缺陷检测/Data/' + date + '/A3138MG000_7K01D72PAK00210/' + num, 'D:/Using AI/烟弹管缺陷检测/Data/' + date + '/pic/210/'+num)
        # cut_img('D:/Using AI/烟弹管缺陷检测/Data/' + date + '/A3138MG000_7K01D72PAK00001/' + num, 'D:/Using AI/烟弹管缺陷检测/Data/' + date + '/pic/001/'+num)
        # detect_YD('D:/Using AI/烟弹管缺陷检测/Data/' + date + '/pic/210/'+num, 'runs/result/' + date, '210/'+num, 0.1, 'weights/best_210.pt')
        # detect_YD('D:/Using AI/烟弹管缺陷检测/Data/' + date + '/pic/001/'+num, 'runs/result/' + date, '001/'+num, 0.3, 'weights/best_001.pt')
        # save_to_excel('C:/代码/烟弹管项目/yolov5-YanDan/runs/result/' + date + '/001/'+num,
        #               'C:/代码/烟弹管项目/yolov5-YanDan/runs/result/' + date + '/210/'+num,
        #               'D:/Using AI/烟弹管缺陷检测/Data/' + date + '/A3138MG000_7K01D72PAK00001/'+num,
        #               'D:/Using AI/烟弹管缺陷检测/Data/' + date + '/A3138MG000_7K01D72PAK00210/' + num,
        #               'D:/Using AI/烟弹管缺陷检测/Data/'+date+'/'+num)
    # 所有excel总到一起
    # mix_excel(date)
    # 拼接并输出错误图像
    # output_wrong_img(date)
    # 错误统计
    # error_stat('C:/代码/烟弹管项目/yolov5-YanDan/runs/result/'+date+'/001', 'D:/Using AI/烟弹管缺陷检测/Data/' + date + '/A3138MG000_7K01D72PAK00001')
    # error_stat('C:/代码/烟弹管项目/yolov5-YanDan/runs/result/'+date+'/210', 'D:/Using AI/烟弹管缺陷检测/Data/' + date + '/A3138MG000_7K01D72PAK00210')



    # -------------
    # 训练前准备
    # -------------

    # dates = ['4.21','4.22','4.23','4.25','4.26','4.27','5.4','5.6','5.11] # 选择训练使用的数据
    # for date in dates:
    #     cut_img('D:/Using AI/烟弹管缺陷检测/Data/'+date+'/A3138MG000_7K01D72PAK00001/1', 'C:/代码/烟弹管项目/yolov5-YanDan/YD_data/images/all_001')
    #     xml_convert('D:/Using AI/烟弹管缺陷检测/Data/'+date+'/A3138MG000_7K01D72PAK00001/1', 'C:/代码/烟弹管项目/yolov5-YanDan/YD_data/labels/all_001', 'D:/Using AI/烟弹管缺陷检测/Data/'+date+'/A3138MG000_7K01D72PAK00001/1')
    #     cut_img('D:/Using AI/烟弹管缺陷检测/Data/'+date+'/A3138MG000_7K01D72PAK00210/1', 'C:/代码/烟弹管项目/yolov5-YanDan/YD_data/images/all_210')
    #     xml_convert('D:/Using AI/烟弹管缺陷检测/Data/'+date+'/A3138MG000_7K01D72PAK00210/1', 'C:/代码/烟弹管项目/yolov5-YanDan/YD_data/labels/all_210', 'D:/Using AI/烟弹管缺陷检测/Data/'+date+'/A3138MG000_7K01D72PAK00210/1')
    # for filename in os.listdir('C:/代码/烟弹管项目/yolov5-YanDan/YD_data/images/all_210'):
    #     r = random.random()
    #     image = cv2.imdecode(np.fromfile('C:/代码/烟弹管项目/yolov5-YanDan/YD_data/images/all_210/'+filename, dtype=np.uint8), -1)
    #     try:
    #         label = open('C:/代码/烟弹管项目/yolov5-YanDan/YD_data/labels/all_210/'+filename[:-4]+'.txt')
    #         if r < 0.9:
    #             cv2.imencode('.bmp', image)[1].tofile('C:/代码/烟弹管项目/yolov5-YanDan/YD_data/images/train/' + filename)
    #             txt = open('C:/代码/烟弹管项目/yolov5-YanDan/YD_data/labels/train/' + filename[:-4] + '.txt', 'w')
    #             txt.write(label.read())
    #         else:
    #             cv2.imencode('.bmp', image)[1].tofile('C:/代码/烟弹管项目/yolov5-YanDan/YD_data/images/val/' + filename)
    #             txt = open('C:/代码/烟弹管项目/yolov5-YanDan/YD_data/labels/val/' + filename[:-4] + '.txt', 'w')
    #             txt.write(label.read())
    #     except:
    #         if r < 0.9:
    #             cv2.imencode('.bmp', image)[1].tofile('C:/代码/烟弹管项目/yolov5-YanDan/YD_data/images/train/' + filename)
    #         else:
    #             cv2.imencode('.bmp', image)[1].tofile('C:/代码/烟弹管项目/yolov5-YanDan/YD_data/images/val/' + filename)


    # -------------
    # 测试有标注框的数据
    # -------------

    # date = '5.27'
    # for num in range(1, 4):
    #     num = str(num)
    #     cut_img('D:/Using AI/烟弹管缺陷检测/Data/' + date + '/A3138MG000_7K01D72PAK00210/' + num, 'D:/Using AI/烟弹管缺陷检测/Data/' + date + '/pic/210/'+num)
    #     cut_img('D:/Using AI/烟弹管缺陷检测/Data/' + date + '/A3138MG000_7K01D72PAK00001/' + num, 'D:/Using AI/烟弹管缺陷检测/Data/' + date + '/pic/001/'+num)
    #     xml_convert('D:/Using AI/烟弹管缺陷检测/Data/' + date + '/A3138MG000_7K01D72PAK00210/'+num,
    #                 'D:/Using AI/烟弹管缺陷检测/Data/' + date + '/labels/210/'+num,
    #                 'D:/Using AI/烟弹管缺陷检测/Data/' + date + '/A3138MG000_7K01D72PAK00210/'+num)
    #     xml_convert('D:/Using AI/烟弹管缺陷检测/Data/' + date + '/A3138MG000_7K01D72PAK00001/' + num,
    #                 'D:/Using AI/烟弹管缺陷检测/Data/' + date + '/labels/001/' + num,
    #                 'D:/Using AI/烟弹管缺陷检测/Data/' + date + '/A3138MG000_7K01D72PAK00001/' + num)
    #     detect_YD('D:/Using AI/烟弹管缺陷检测/Data/' + date + '/pic/210/'+num, 'runs/result/' + date, '210/'+num, 0.1, 'weights/best_210.pt')
    #     detect_YD('D:/Using AI/烟弹管缺陷检测/Data/' + date + '/pic/001/'+num, 'runs/result/' + date, '001/'+num, 0.1, 'weights/best_001.pt')
    #     save_to_excel('C:/代码/烟弹管项目/yolov5-YanDan/runs/result/' + date + '/001/'+num,
    #                   'C:/代码/烟弹管项目/yolov5-YanDan/runs/result/' + date + '/210/'+num,
    #                   'D:/Using AI/烟弹管缺陷检测/Data/' + date + '/labels/001/'+num,
    #                   'D:/Using AI/烟弹管缺陷检测/Data/' + date + '/labels/210/'+num,
    #                   'D:/Using AI/烟弹管缺陷检测/Data/'+date+'/'+num)
    ## 所有excel总到一起
    # mix_excel(date)
    ## 拼接并输出错误图像
    # output_wrong_img(date)
    ## 统计错误类型并画柱状图
    # error_stat(label_directory_pre='C:/代码/烟弹管项目/yolov5-YanDan/runs/result/'+date+'/210',
    #            label_directory_true='D:/Using AI/烟弹管缺陷检测/Data/'+date+'/A3138MG000_7K01D72PAK00210',
    #            save_path='D:/Using AI/烟弹管缺陷检测/Data/'+date+'/stat_210.png')
    # error_stat(label_directory_pre='C:/代码/烟弹管项目/yolov5-YanDan/runs/result/'+date+'/001',
    #            label_directory_true='D:/Using AI/烟弹管缺陷检测/Data/'+date+'/A3138MG000_7K01D72PAK00001',
    #            save_path='D:/Using AI/烟弹管缺陷检测/Data/'+date+'/stat_001.png')


    # -------------
    # 测试4.1组织结构的数据
    # -------------

    # for num in os.listdir('D:/Using AI/烟弹管缺陷检测/Data/'+date+'/OK/上相机'):
    #     cut_img('D:/Using AI/烟弹管缺陷检测/Data/' + date + '/OK/上相机/' + num, 'D:/Using AI/烟弹管缺陷检测/Data/' + date + '/pic/OK/210/'+num)
    #     cut_img('D:/Using AI/烟弹管缺陷检测/Data/' + date + '/OK/下相机/' + num, 'D:/Using AI/烟弹管缺陷检测/Data/' + date + '/pic/OK/001/'+num)
    #     detect_YD('D:/Using AI/烟弹管缺陷检测/Data/' + date + '/pic/OK/210/'+num, 'runs/result/' + date, 'OK/210/'+num, 0.1, 'weights/best_210.pt')
    #     detect_YD('D:/Using AI/烟弹管缺陷检测/Data/' + date + '/pic/OK/001/'+num, 'runs/result/' + date, 'OK/001/'+num, 0.1, 'weights/best_001.pt')
    #     save_to_excel('C:/代码/烟弹管项目/yolov5-YanDan/runs/result/' + date + '/OK/001/'+num,
    #                   'C:/代码/烟弹管项目/yolov5-YanDan/runs/result/' + date + '/OK/210/'+num,
    #                   'D:/Using AI/烟弹管缺陷检测/Data/' + date + '/OK/下相机/'+num,
    #                   'D:/Using AI/烟弹管缺陷检测/Data/' + date + '/OK/上相机/'+num,
    #                   'D:/Using AI/烟弹管缺陷检测/Data/'+date+'/OK/'+num,
    #                   label=None)
    # for num in os.listdir('D:/Using AI/烟弹管缺陷检测/Data/' + date + '/NG/上相机'):
        # cut_img('D:/Using AI/烟弹管缺陷检测/Data/' + date + '/NG/上相机/' + num,
        #         'D:/Using AI/烟弹管缺陷检测/Data/' + date + '/pic/NG/210/' + num)
        # cut_img('D:/Using AI/烟弹管缺陷检测/Data/' + date + '/NG/下相机/' + num,
        #         'D:/Using AI/烟弹管缺陷检测/Data/' + date + '/pic/NG/001/' + num)
        # detect_YD('D:/Using AI/烟弹管缺陷检测/Data/' + date + '/pic/NG/210/' + num, 'runs/result/' + date, 'NG/210/' + num, 0.1, 'weights/best_210.pt')
        # detect_YD('D:/Using AI/烟弹管缺陷检测/Data/' + date + '/pic/NG/001/' + num, 'runs/result/' + date, 'NG/001/' + num, 0.1, 'weights/best_001.pt')
        # save_to_excel('C:/代码/烟弹管项目/yolov5-YanDan/runs/result/' + date + '/NG/001/' + num,
        #               'C:/代码/烟弹管项目/yolov5-YanDan/runs/result/' + date + '/NG/210/' + num,
        #               'D:/Using AI/烟弹管缺陷检测/Data/' + date + '/NG/下相机/' + num,
        #               'D:/Using AI/烟弹管缺陷检测/Data/' + date + '/NG/上相机/' + num,
        #               'D:/Using AI/烟弹管缺陷检测/Data/' + date + '/NG/' + num,
        #               label='NG')
    ## 错误统计
    # error_stat('C:/代码/烟弹管项目/yolov5-YanDan/runs/result/'+date+'/001', 'D:/Using AI/烟弹管缺陷检测/Data/' + date + '/A3138MG000_7K01D72PAK00001')
    # error_stat('C:/代码/烟弹管项目/yolov5-YanDan/runs/result/'+date+'/210', 'D:/Using AI/烟弹管缺陷检测/Data/' + date + '/A3138MG000_7K01D72PAK00210')











